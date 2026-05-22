"""
SFP Report Generator  —  SFP-02
Gerado automaticamente ao final de cada execucao do sfp_analyzer.py.
Pode ser re-executado de forma independente sem precisar rodar a LLM:

    python src/report/generate_report.py

Entrada : output/sfp/sfp_consolidated.json
Saida   : output/sfp_report.xlsx

Abas:
  1. Resumo           — painel por repositorio com o funil de classificacao
  2. Elementos Contados — todos os FD e EP com fonte e razao (auditavel)
  3. Auditoria LLM    — somente itens decididos pela LLM, incluindo ignorados
"""

import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ─────────────────────────────────────────
# Paleta e estilos
# ─────────────────────────────────────────
HDR_BG   = "1E3A5F"   # azul escuro — cabecalho principal
HDR_FG   = "FFFFFF"
SUB_BG   = "2E75B6"   # azul medio — cabecalho de secao
SUB_FG   = "FFFFFF"
FD_BG    = "E6F1FB"   # azul claro  — linhas FD
FD_FG    = "0C447C"
EP_BG    = "EAF3DE"   # verde claro — linhas EP
EP_FG    = "3B6D11"
IGN_BG   = "F5F5F5"   # cinza claro — ignorados
IGN_FG   = "666666"
LLM_BG   = "FAEEDA"   # ambar claro — fonte LLM
LLM_FG   = "633806"
PRE_BG   = "F0F0F0"   # cinza muito claro — fonte pre_classifier
PRE_FG   = "444444"
TOT_BG   = "F5F8FC"   # total / alternado
BORDER_C = "CCCCCC"

thin  = Side(style="thin",   color=BORDER_C)
thick = Side(style="medium", color=HDR_BG)
brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
brd_t = Border(left=thin, right=thin, top=thick, bottom=thin)

EXT_LANG = {
    ".py":   "Python",
    ".java": "Java",
    ".cs":   "C#",
    ".ts":   "TypeScript",
    ".tsx":  "TypeScript",
    ".js":   "JavaScript",
    ".jsx":  "JavaScript",
    ".kt":   "Kotlin",
}


# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────
def lang_from_file(filepath: str) -> str:
    ext = Path(filepath.replace("\\", "/")).suffix.lower()
    return EXT_LANG.get(ext, "")


def short_file(filepath: str) -> str:
    """Remove o prefixo do repo e normaliza separadores."""
    return filepath.replace("\\", "/").lstrip("/")


def _cell(ws, row, col, value="", bold=False, fg="000000", bg=None,
          wrap=False, halign="left", size=9, border=None, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=size, italic=italic)
    c.alignment = Alignment(horizontal=halign, vertical="center",
                            wrap_text=wrap)
    c.border    = border or brd
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def _header_row(ws, row, labels, col_widths, bg=HDR_BG, fg=HDR_FG, height=22):
    ws.row_dimensions[row].height = height
    for col, (label, width) in enumerate(zip(labels, col_widths), 1):
        _cell(ws, row, col, label, bold=True, fg=fg, bg=bg,
              halign="center", size=10)
        ws.column_dimensions[get_column_letter(col)].width = width


def _title_row(ws, row, text, ncols, height=28):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, color=HDR_FG, size=12)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = brd
    ws.row_dimensions[row].height = height


def _subtitle_row(ws, row, text, ncols, height=16):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", color="888888", size=8, italic=True)
    c.fill      = PatternFill("solid", fgColor="F5F8FC")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = brd
    ws.row_dimensions[row].height = height


# ─────────────────────────────────────────
# Aba 1 — Resumo
# ─────────────────────────────────────────
def _build_summary(ws, repos):
    COLS = [
        "Repositorio",
        "FD (auto)", "FD (LLM)", "FD Total",
        "EP (auto)", "EP (LLM)", "EP Total",
        "Enviados LLM", "Ignorados LLM",
        "SFP Total",
    ]
    WIDTHS = [30, 10, 10, 10, 10, 10, 10, 14, 14, 12]
    N = len(COLS)

    _title_row(ws, 1, "SFP PoC  —  Resumo por Repositorio", N)
    _subtitle_row(ws, 2,
                  "Funil de classificacao: pre-classificador + LLM  |  "
                  "Atualizado automaticamente a cada execucao do sfp_analyzer.py", N)
    _header_row(ws, 3, COLS, WIDTHS)
    ws.freeze_panes = "A4"

    total_fd = total_ep = total_sfp = 0
    auto_fd_t = auto_ep_t = llm_fd_t = llm_ep_t = llm_sent_t = llm_ign_t = 0

    for ri, repo in enumerate(repos, 4):
        ws.row_dimensions[ri].height = 18
        bg = "FFFFFF" if ri % 2 == 0 else TOT_BG

        pre = repo.get("pre_classification", {})
        llm = repo.get("llm_classification", {})
        cnt = repo.get("sfp_count", {})

        auto_fd  = pre.get("data_functions", 0)
        auto_ep  = pre.get("elementary_processes", 0)
        llm_fd   = llm.get("data_functions", 0)
        llm_ep   = llm.get("elementary_processes", 0)
        sent     = llm.get("sent", 0)
        ignored  = llm.get("ignored", 0)
        fd_total = cnt.get("data_functions", 0)
        ep_total = cnt.get("elementary_processes", 0)
        sfp      = cnt.get("total", 0)

        total_fd  += fd_total
        total_ep  += ep_total
        total_sfp += sfp
        auto_fd_t += auto_fd; auto_ep_t += auto_ep
        llm_fd_t  += llm_fd;  llm_ep_t  += llm_ep
        llm_sent_t += sent;   llm_ign_t += ignored

        vals = [repo["repository"], auto_fd, llm_fd, fd_total,
                auto_ep, llm_ep, ep_total, sent, ignored, sfp]

        for ci, v in enumerate(vals, 1):
            halign = "left" if ci == 1 else "center"
            bold   = ci in (4, 7, 10)
            fg     = FD_FG if ci == 4 else (EP_FG if ci == 7 else
                     ("1E3A5F" if ci == 10 else "333333"))
            _cell(ws, ri, ci, v, bold=bold, fg=fg, bg=bg, halign=halign)

    # Linha de totais
    tr = len(repos) + 4
    ws.row_dimensions[tr].height = 20
    totals = ["TOTAL", auto_fd_t, llm_fd_t, total_fd,
              auto_ep_t, llm_ep_t, total_ep, llm_sent_t, llm_ign_t, total_sfp]
    for ci, v in enumerate(totals, 1):
        halign = "left" if ci == 1 else "center"
        _cell(ws, tr, ci, v, bold=True, fg=HDR_FG, bg=HDR_BG,
              halign=halign, size=10, border=brd_t)

    # Legenda
    lr = tr + 2
    ws.row_dimensions[lr].height = 14
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=N)
    lc = ws.cell(row=lr, column=1,
                 value="FD = Funcao de Dados  |  EP = Processo Elementar  |  "
                       "auto = pre-classificado pelo extrator  |  "
                       "LLM = classificado pelo Azure OpenAI")
    lc.font      = Font(name="Arial", size=8, color="888888", italic=True)
    lc.alignment = Alignment(horizontal="left", vertical="center")
    lc.border    = brd


# ─────────────────────────────────────────
# Aba 2 — Elementos Contados
# ─────────────────────────────────────────
def _build_elements(ws, repos):
    COLS   = ["Repositorio", "Tipo", "Nome", "Arquivo", "Linguagem", "Fonte", "Razao"]
    WIDTHS = [28, 6, 32, 52, 13, 17, 68]
    N = len(COLS)

    _title_row(ws, 1, "SFP PoC  —  Elementos Contados (FD + EP)", N)
    _subtitle_row(ws, 2,
                  "Todos os simbolos incluidos na contagem SFP final  |  "
                  "Use filtros para auditar por repositorio, tipo ou fonte", N)
    _header_row(ws, 3, COLS, WIDTHS)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(N)}3"

    row = 4
    for repo in repos:
        name = repo["repository"]
        items = (
            [(x, "FD") for x in repo.get("data_functions", [])] +
            [(x, "EP") for x in repo.get("elementary_processes", [])]
        )
        for item, tipo in items:
            ws.row_dimensions[row].height = 42
            bg     = FD_BG  if tipo == "FD" else EP_BG
            tipo_fg = FD_FG if tipo == "FD" else EP_FG
            src    = item.get("source", "")
            reason = item.get("reason", "")
            if not reason and src == "pre_classifier":
                reason = "(razao disponivel apos proxima execucao do extrator)"

            src_label = "pre-classificador" if src == "pre_classifier" else "LLM"
            src_bg    = PRE_BG if src == "pre_classifier" else LLM_BG
            src_fg    = PRE_FG if src == "pre_classifier" else LLM_FG

            filepath = item.get("file", "")

            _cell(ws, row, 1, name,                   fg="1E3A5F", bg=bg, size=9)
            _cell(ws, row, 2, tipo,   bold=True,       fg=tipo_fg, bg=bg, halign="center")
            _cell(ws, row, 3, item.get("name", ""),    fg="111111", bg=bg, bold=True)
            _cell(ws, row, 4, short_file(filepath),    fg="555555", bg=bg, wrap=True, size=8)
            _cell(ws, row, 5, lang_from_file(filepath),fg="333333", bg=bg, halign="center")
            _cell(ws, row, 6, src_label, bold=True,    fg=src_fg,  bg=src_bg, halign="center")
            _cell(ws, row, 7, reason,                  fg="333333", bg=bg, wrap=True, size=8)
            row += 1


# ─────────────────────────────────────────
# Aba 3 — Auditoria LLM
# ─────────────────────────────────────────
def _build_llm_audit(ws, repos):
    COLS   = ["Repositorio", "Classificacao LLM", "Nome", "Arquivo",
              "Linguagem", "Razao da LLM"]
    WIDTHS = [28, 20, 32, 52, 13, 72]
    N = len(COLS)

    _title_row(ws, 1, "SFP PoC  —  Auditoria das Decisoes da LLM", N)
    _subtitle_row(ws, 2,
                  "Somente itens enviados ao Azure OpenAI  |  "
                  "Inclui o que foi contado E o que foi ignorado pela LLM", N)
    _header_row(ws, 3, COLS, WIDTHS)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(N)}3"

    CLF_STYLE = {
        "data_function":      ("FD",       FD_BG,  FD_FG),
        "elementary_process": ("EP",       EP_BG,  EP_FG),
        "ignore":             ("ignorado", IGN_BG, IGN_FG),
    }

    row = 4
    for repo in repos:
        name = repo["repository"]

        # Itens contados pela LLM
        llm_items = [
            (x, "data_function")
            for x in repo.get("data_functions", [])
            if x.get("source") == "llm"
        ] + [
            (x, "elementary_process")
            for x in repo.get("elementary_processes", [])
            if x.get("source") == "llm"
        ] + [
            (x, "ignore")
            for x in repo.get("ignored_by_llm", [])
        ]

        for item, clf in llm_items:
            ws.row_dimensions[row].height = 52
            label, bg, fg = CLF_STYLE.get(clf, ("?", "FFFFFF", "000000"))
            filepath = item.get("file", "")

            _cell(ws, row, 1, name,                       fg="1E3A5F",  bg=bg)
            _cell(ws, row, 2, label, bold=True,           fg=fg,        bg=bg, halign="center")
            _cell(ws, row, 3, item.get("name", ""),       fg="111111",  bg=bg, bold=True)
            _cell(ws, row, 4, short_file(filepath),       fg="555555",  bg=bg, wrap=True, size=8)
            _cell(ws, row, 5, lang_from_file(filepath),   fg="333333",  bg=bg, halign="center")
            _cell(ws, row, 6, item.get("reason", ""),     fg="333333",  bg=bg, wrap=True, size=8)
            row += 1

    if row == 4:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=N)
        _cell(ws, 4, 1, "Nenhum item foi enviado a LLM nesta execucao.",
              fg="888888", italic=True)


# ─────────────────────────────────────────
# Funcao principal
# ─────────────────────────────────────────
def generate(sfp_dir: str, output_path: str) -> None:
    consolidated = Path(sfp_dir) / "sfp_consolidated.json"
    if not consolidated.exists():
        print(f"   ⚠️  sfp_consolidated.json nao encontrado em: {sfp_dir}")
        return

    with open(consolidated, encoding="utf-8") as f:
        repos = json.load(f)

    # Filtra repositorios sem contagem (edge cases, frontends)
    repos = [r for r in repos if r.get("sfp_count", {}).get("total", 0) > 0]

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumo"
    _build_summary(ws1, repos)

    ws2 = wb.create_sheet("Elementos Contados")
    _build_elements(ws2, repos)

    ws3 = wb.create_sheet("Auditoria LLM")
    _build_llm_audit(ws3, repos)

    # Seleciona a aba Resumo ao abrir
    wb.active = ws1

    wb.save(output_path)
    print(f"   📊 Relatorio Excel: {output_path}")


# ─────────────────────────────────────────
# Execucao independente
# ─────────────────────────────────────────
if __name__ == "__main__":
    base      = Path(__file__).resolve().parents[2]
    sfp_dir   = base / "output" / "sfp"
    out_file  = base / "output" / "sfp_report.xlsx"
    generate(str(sfp_dir), str(out_file))
